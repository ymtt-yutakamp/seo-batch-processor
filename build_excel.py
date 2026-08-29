#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# データファイルのパス
RESULTS_JSON = Path(__file__).parent / "data" / "results.json"
OUTPUT_XLSX = Path(__file__).parent / "SEO記事AI代行会社リスト.xlsx"

def load_results():
    """results.json からデータを読み込む"""
    if RESULTS_JSON.exists():
        with open(RESULTS_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def build_excel():
    """Excel ファイルを再生成"""
    results = load_results()

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO記事AI代行会社"

    # 行1: 注記
    note_text = "※リスティング広告枠は取得ツールの制限により掲載していません。表示順位は検索API結果の順序であり、実際のGoogle検索結果と異なる可能性があります。"
    ws['A1'] = note_text
    ws['A1'].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells('A1:I1')
    ws.row_dimensions[1].height = 25

    # 行2: ヘッダー
    headers = ["検索キーワード", "リスティングorオーガニック", "表示順位", "会社名", "サービス名", "URL", "特徴", "費用", "サイト種別"]
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True, size=11)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    # 行3以降: データ
    for row_idx, record in enumerate(results, start=3):
        ws.cell(row=row_idx, column=1, value=record.get("検索キーワード", ""))
        ws.cell(row=row_idx, column=2, value=record.get("リスティングorオーガニック", "オーガニック"))
        ws.cell(row=row_idx, column=3, value=record.get("表示順位", ""))
        ws.cell(row=row_idx, column=4, value=record.get("会社名", ""))
        ws.cell(row=row_idx, column=5, value=record.get("サービス名", ""))
        ws.cell(row=row_idx, column=6, value=record.get("URL", ""))
        ws.cell(row=row_idx, column=7, value=record.get("特徴", ""))
        ws.cell(row=row_idx, column=8, value=record.get("費用", ""))
        ws.cell(row=row_idx, column=9, value=record.get("サイト種別", ""))

        # テキスト折り返し＆行高さ
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = None  # 自動高さ

    # 列幅の自動調整
    col_widths = [20, 15, 10, 15, 15, 35, 25, 20, 22]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Excel 保存
    wb.save(OUTPUT_XLSX)
    print(f"✓ Excel 保存: {OUTPUT_XLSX}")
    print(f"  総行数: {len(results)} 件")

    return len(results)

if __name__ == "__main__":
    try:
        count = build_excel()
        sys.exit(0)
    except Exception as e:
        print(f"✗ エラー: {e}", file=sys.stderr)
        sys.exit(1)
